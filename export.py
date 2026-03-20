"""
export.py  —  Gera relatório .xlsx com resumo e layouts por chapa
"""

import io
import math
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.patches import FancyArrowPatch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter


def _hex_rgb(h):
    h = h.lstrip('#')
    return tuple(int(h[i:i+2], 16)/255 for i in (0, 2, 4))

def _darker(h, f=0.6):
    r,g,b = _hex_rgb(h)
    return (r*f, g*f, b*f)


def _render_chapa(chapa_dict, sheet_w, sheet_h, margin_x, margin_y, idx, total):
    fig_w = 9
    fig_h = fig_w * (sheet_h / sheet_w)
    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    ax.set_facecolor('#E0E0E0')
    fig.patch.set_facecolor('#FFFFFF')

    # Fundo da chapa
    ax.add_patch(mpatches.Rectangle(
        (0, 0), sheet_w, sheet_h,
        lw=2, ec='#1F497D', fc='#F5F5F5', zorder=1))

    # Área útil
    ax.add_patch(mpatches.Rectangle(
        (margin_x, margin_y),
        sheet_w - 2*margin_x, sheet_h - 2*margin_y,
        lw=1, ec='#AAAAAA', fc='#FAFAFA', ls='--', zorder=2))

    # Peças — polígonos reais com suporte a furos
    from matplotlib.patches import PathPatch
    from matplotlib.path import Path as MPath

    for p in chapa_dict['pecas']:
        cor   = p['cor']
        borda = _darker(cor)
        poly_coords = p.get('poly_coords', [])
        hole_coords = p.get('hole_coords', [])

        if poly_coords and len(poly_coords) >= 3:
            verts = [(x, y) for x, y in poly_coords]
            codes = ([MPath.MOVETO] +
                     [MPath.LINETO] * (len(verts) - 2) +
                     [MPath.CLOSEPOLY])
            for hole in hole_coords:
                if len(hole) >= 3:
                    verts += [(x, y) for x, y in hole]
                    codes += ([MPath.MOVETO] +
                               [MPath.LINETO] * (len(hole) - 2) +
                               [MPath.CLOSEPOLY])
            patch = PathPatch(MPath(verts, codes),
                              lw=1.2, ec=borda,
                              fc=(*_hex_rgb(cor), 0.85), zorder=3)
            ax.add_patch(patch)
            xs = [v[0] for v in poly_coords]
            ys = [v[1] for v in poly_coords]
            cx = (min(xs) + max(xs)) / 2
            cy = (min(ys) + max(ys)) / 2
        else:
            ax.add_patch(mpatches.Rectangle(
                (p['x'], p['y']), p['w'], p['h'],
                lw=1.2, ec=borda, fc=(*_hex_rgb(cor), 0.85), zorder=3))
            cx = p['x'] + p['w'] / 2
            cy = p['y'] + p['h'] / 2

        fs = max(5, min(9, p['w']/60, p['h']/20))
        ang_str = f"\n↺{p.get('angulo',0):.0f}°" if p.get('angulo') else ''
        ax.text(cx, cy,
                f"{p['nome']}\n{p['label']}{ang_str}",
                ha='center', va='center', fontsize=fs,
                fontweight='bold', color='white', zorder=4,
                bbox=dict(boxstyle='round,pad=0.15', fc='none', ec='none'))

    # Cota X
    ax.annotate('', xy=(sheet_w, -sheet_h*0.02),
                xytext=(0, -sheet_h*0.02),
                arrowprops=dict(arrowstyle='<->', color='#444', lw=1))
    ax.text(sheet_w/2, -sheet_h*0.035, f'{sheet_w:.0f} mm',
            ha='center', va='top', fontsize=8, color='#444')

    # Cota Y
    ax.annotate('', xy=(sheet_w*1.02, sheet_h),
                xytext=(sheet_w*1.02, 0),
                arrowprops=dict(arrowstyle='<->', color='#444', lw=1))
    ax.text(sheet_w*1.025, sheet_h/2, f'{sheet_h:.0f} mm',
            ha='left', va='center', fontsize=8, color='#444', rotation=90)

    ax.set_xlim(-sheet_w*0.02, sheet_w*1.08)
    ax.set_ylim(-sheet_h*0.06, sheet_h*1.04)
    ax.set_aspect('equal')
    ax.set_title(
        f'Chapa {idx}/{total}  —  {chapa_dict["n_pecas"]} peças  '
        f'|  Aproveitamento: {chapa_dict["aproveitamento"]:.1f}%',
        fontsize=11, fontweight='bold', color='#1F497D', pad=8)
    ax.axis('off')
    plt.tight_layout(pad=0.3)

    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=130, bbox_inches='tight', facecolor='white')
    plt.close(fig)
    buf.seek(0)
    return buf


def gerar(chapas_dict, pecas_cfg, config, output_path):
    wb = Workbook()

    # ── Estilos ──────────────────────────────────────────────
    BLU  = PatternFill("solid", fgColor="FF1F497D")
    BLUL = PatternFill("solid", fgColor="FFDCE6F1")
    GRY  = PatternFill("solid", fgColor="FFE8E8E8")
    WHT  = PatternFill("solid", fgColor="FFFFFFFF")
    GOLD = PatternFill("solid", fgColor="FF7F6000")
    GLDL = PatternFill("solid", fgColor="FFFFF2CC")
    thin = Side(style='thin', color='FFCCCCCC')
    TB   = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hc(cell, val, fill=BLU, sz=10, bold=True, clr='FFFFFFFF', align='center'):
        cell.value = val
        cell.fill  = fill
        cell.font  = Font(name='Arial', size=sz, bold=bold, color=clr)
        cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
        cell.border = TB

    def vc(cell, val, fmt=None, fill=WHT, sz=10, bold=False, clr='FF000000', align='center'):
        cell.value = val
        cell.fill  = fill
        cell.font  = Font(name='Arial', size=sz, bold=bold, color=clr)
        cell.alignment = Alignment(horizontal=align, vertical='center')
        cell.border = TB
        if fmt: cell.number_format = fmt

    # ── ABA RESUMO ───────────────────────────────────────────
    ws = wb.active
    ws.title = 'RESUMO'
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = '1F497D'

    for col, w in zip('ABCDE', [28,16,16,16,16]):
        ws.column_dimensions[col].width = w

    ws.row_dimensions[1].height = 30
    ws.merge_cells('A1:E1')
    hc(ws['A1'], f'NESTING 2D — {config.get("nome","Projeto")}', sz=13)

    # Parâmetros
    ws.merge_cells('A3:E3')
    hc(ws['A3'], 'PARÂMETROS DA CHAPA E PROCESSO')
    pars = [
        ('Material',           config.get('material','')),
        ('Espessura (mm)',      config.get('espessura','')),
        ('Chapa X × Y (mm)',   f'{config.get("sheet_w","")} × {config.get("sheet_h","")}'),
        ('Gap (mm)',            config.get('gap','')),
        ('Margem X / Y (mm)',  f'{config.get("margin_x","")} / {config.get("margin_y","")}'),
        ('Preço MP (R$/kg)',    config.get('preco_kg',0)),
    ]
    for i,(lbl,val) in enumerate(pars):
        r = 4+i
        ws.row_dimensions[r].height = 15
        hc(ws.cell(r,1), lbl, fill=BLUL, clr='FF1F497D', bold=True, align='left')
        vc(ws.cell(r,2), val, bold=True, clr='FF1F497D')

    # Resultados
    n_chapas   = len(chapas_dict)
    aprov_med  = sum(c['aproveitamento'] for c in chapas_dict)/n_chapas if n_chapas else 0
    sheet_w    = config.get('sheet_w', 1200)
    sheet_h    = config.get('sheet_h', 2000)
    esp        = config.get('espessura', 1.5)
    rho        = config.get('rho', 7850)
    preco_kg   = config.get('preco_kg', 0)
    pintura    = config.get('pintura_m2', 0)
    icms       = config.get('icms_pct', 18)
    outros     = config.get('outros_pct', 9.25)
    lucro      = config.get('lucro_pct', 30)

    massa_chapa  = (sheet_w/1000)*(sheet_h/1000)*(esp/1000)*rho
    massa_total  = massa_chapa * n_chapas
    custo_mp     = massa_chapa * preco_kg * n_chapas

    rr = 4+len(pars)+2
    ws.merge_cells(f'A{rr-1}:E{rr-1}')
    hc(ws[f'A{rr-1}'], 'RESULTADOS', fill=GOLD)
    res = [
        ('Chapas necessárias',   n_chapas,        '0'),
        ('Aproveitamento médio', aprov_med/100,   '0.0%'),
        ('Peso total (kg)',       round(massa_total, 2), '#,##0.00'),
        ('Custo MP total (R$)',  custo_mp,        'R$ #,##0.00'),
    ]
    for i,(lbl,val,fmt) in enumerate(res):
        r = rr+i
        ws.row_dimensions[r].height = 16
        hc(ws.cell(r,1), lbl, fill=GLDL, clr='FF7F6000', bold=True, align='left')
        vc(ws.cell(r,2), val, fmt=fmt, bold=True, clr='FF7F6000')

    # Tabela de peças
    tr = rr+len(res)+2
    ws.merge_cells(f'A{tr-1}:E{tr-1}')
    hc(ws[f'A{tr-1}'], 'PEÇAS')
    for j,h in enumerate(['Peça','Largura (mm)','Altura (mm)','Qtd. sol.','Qtd. aloc.']):
        hc(ws.cell(tr,j+1), h, fill=PatternFill("solid",fgColor="FF366092"), sz=9)

    aloc = {}
    for c in chapas_dict:
        for p in c['pecas']:
            aloc[p['nome']] = aloc.get(p['nome'],0)+1

    for i,pc in enumerate(pecas_cfg):
        r = tr+1+i
        ws.row_dimensions[r].height = 15
        ch = pc.get('cor','#4472C4').lstrip('#').upper()
        fp = PatternFill("solid", fgColor=f'FF{ch}')
        for j,val in enumerate([pc['nome'], pc.get('largura',''), pc.get('altura',''),
                                  pc['quantidade'], aloc.get(pc['nome'],0)]):
            vc(ws.cell(r,j+1), val, fill=fp if j==0 else GRY if i%2==0 else WHT,
               clr='FFFFFFFF' if j==0 else 'FF000000', sz=9)

    # ── ABAS POR CHAPA ────────────────────────────────────────
    for ch in chapas_dict:
        wsc = wb.create_sheet(f'Chapa {ch["indice"]}')
        wsc.sheet_view.showGridLines = False
        wsc.sheet_properties.tabColor = '366092'
        for col,w in zip('ABCDE',[22,14,14,14,14]):
            wsc.column_dimensions[col].width = w

        wsc.row_dimensions[1].height = 22
        wsc.merge_cells('A1:E1')
        hc(wsc['A1'],
           f'Chapa {ch["indice"]}/{n_chapas}  —  {ch["n_pecas"]} peças  '
           f'|  Aproveitamento: {ch["aproveitamento"]:.1f}%', sz=11)

        buf = _render_chapa(ch, sheet_w, sheet_h,
                             config.get('margin_x',0), config.get('margin_y',0),
                             ch['indice'], n_chapas)
        img = XLImage(buf)
        img.width  = 500
        img.height = int(500 * sheet_h / sheet_w)
        wsc.add_image(img, 'A3')

        tr2 = 3 + img.height//14 + 2
        wsc.merge_cells(f'A{tr2}:E{tr2}')
        hc(wsc[f'A{tr2}'], 'Peças nesta chapa')
        for j,h in enumerate(['Peça','X','Y','L×A (mm)','Girada']):
            hc(wsc.cell(tr2+1,j+1), h,
               fill=PatternFill("solid",fgColor="FF366092"), sz=9)
        for i,p in enumerate(ch['pecas']):
            r = tr2+2+i
            ch2 = p['cor'].lstrip('#').upper()
            fp  = PatternFill("solid", fgColor=f'FF{ch2}')
            for j,val in enumerate([p['nome'], round(p['x'],1), round(p['y'],1),
                                      p['label'], 'Sim' if p['girada'] else 'Não']):
                vc(wsc.cell(r,j+1), val,
                   fill=fp if j==0 else GRY if i%2==0 else WHT,
                   clr='FFFFFFFF' if j==0 else 'FF000000', sz=9)

    wb.save(output_path)
    return output_path
